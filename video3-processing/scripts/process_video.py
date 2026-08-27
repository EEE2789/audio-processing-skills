#!/usr/bin/env python3
"""
Video3 字幕｜行情视频字幕版（1.1x + 2秒简体封面 + 烧录字幕）

基于原始视频生成 Video3 成品视频。
为带字幕版本，用于公开视频平台发布。

输入：原始视频 + 字幕文件 + 简体封面
输出：/Users/ai/Documents/video_pipeline/2output/3字幕_MMDD.mp4
"""

import sys
import os
import subprocess
import re
import tempfile
import shutil
import fcntl
from datetime import datetime
from openpyxl import load_workbook

# ====== 固定配置 ======
FFMPEG_PATH = "/opt/homebrew/bin/ffmpeg"
INPUT_DIR = "/Users/ai/Documents/video_pipeline/1input"
SUBTITLE_DIR = "/Users/ai/Documents/video_pipeline/3daily"
COVER_DIR = "/Users/ai/Documents/video_pipeline/3daily/covers"
OUTPUT_DIR = "/Users/ai/Documents/video_pipeline/2output"
METADATA_EXCEL = "/Users/ai/Documents/video_pipeline/2output/视频自动上传.xlsx"

# 编码参数（固定）
VIDEO_CODEC = os.getenv("VIDEO_CODEC", "h264_videotoolbox")
VIDEO_BITRATE = os.getenv("VIDEO_BITRATE", "8000k")
OUTPUT_WIDTH = int(os.getenv("OUTPUT_WIDTH", "1920"))
PIXEL_FMT = "yuv420p"
PROFILE = "high"
LEVEL = "5.2"
CRF = "15"
PRESET = "veryfast"
FPS = "24"
AUDIO_CODEC = "aac"
AUDIO_BITRATE = "320k"


def acquire_process_lock():
    """防止同一个 Video3 任务被并发启动"""
    lock_path = "/tmp/video_pipeline_video3.lock"
    lock_file = open(lock_path, "w")
    try:
        fcntl.flock(lock_file, fcntl.LOCK_EX | fcntl.LOCK_NB)
    except BlockingIOError:
        raise RuntimeError("❌ Video3 正在生成中，请等待当前任务完成后再运行")
    return lock_file


def validate_ffmpeg():
    """校验 ffmpeg 是否存在"""
    if not os.path.exists(FFMPEG_PATH):
        raise RuntimeError(f"❌ ffmpeg 不存在于指定路径: {FFMPEG_PATH}")


def get_latest_video():
    """自动获取 1input 文件夹中最新的视频文件"""
    if not os.path.exists(INPUT_DIR):
        raise RuntimeError(f"❌ 输入目录不存在: {INPUT_DIR}")

    # 支持的视频格式
    video_extensions = ['.mp4', '.mov', '.avi', '.mkv', '.flv', '.wmv', '.m4v']

    # 获取所有视频文件
    video_files = []
    for file in os.listdir(INPUT_DIR):
        if any(file.lower().endswith(ext) for ext in video_extensions):
            file_path = os.path.join(INPUT_DIR, file)
            video_files.append((file_path, os.path.getmtime(file_path)))

    if not video_files:
        raise RuntimeError(f"❌ 在 {INPUT_DIR} 中未找到视频文件")

    # 按修改时间排序，获取最新的
    video_files.sort(key=lambda x: x[1], reverse=True)
    latest_video = video_files[0][0]

    print(f"📂 自动找到最新视频: {os.path.basename(latest_video)}")
    return latest_video


def get_latest_subtitle():
    """自动获取 3daily 文件夹中最新的简体字幕"""
    if not os.path.exists(SUBTITLE_DIR):
        raise RuntimeError(f"❌ 字幕目录不存在: {SUBTITLE_DIR}")

    # 查找简体字幕
    srt_files = []
    for file in os.listdir(SUBTITLE_DIR):
        if file.startswith('简体') and file.endswith('.srt'):
            file_path = os.path.join(SUBTITLE_DIR, file)
            srt_files.append((file_path, os.path.getmtime(file_path)))

    if not srt_files:
        raise RuntimeError(f"❌ 在 {SUBTITLE_DIR} 中未找到简体字幕 (简体*.srt)")

    # 按修改时间排序，获取最新的
    srt_files.sort(key=lambda x: x[1], reverse=True)
    latest_srt = srt_files[0][0]

    print(f"📂 自动找到最新字幕: {os.path.basename(latest_srt)}")
    return latest_srt


def get_latest_cover():
    """自动获取 covers 文件夹中最新的简体封面"""
    if not os.path.exists(COVER_DIR):
        raise RuntimeError(f"❌ 封面目录不存在: {COVER_DIR}")

    # 查找简体封面
    cover_files = []
    for file in os.listdir(COVER_DIR):
        if 'simplified' in file.lower() and (file.endswith('.png') or file.endswith('.jpg')):
            file_path = os.path.join(COVER_DIR, file)
            cover_files.append((file_path, os.path.getmtime(file_path)))

    if not cover_files:
        # 如果没有简体封面，找任何 PNG/JPG
        for file in os.listdir(COVER_DIR):
            if file.endswith('.png') or file.endswith('.jpg') or file.endswith('.jpeg'):
                file_path = os.path.join(COVER_DIR, file)
                cover_files.append((file_path, os.path.getmtime(file_path)))

    if not cover_files:
        raise RuntimeError(f"❌ 在 {COVER_DIR} 中未找到封面图片")

    # 按修改时间排序，获取最新的
    cover_files.sort(key=lambda x: x[1], reverse=True)
    latest_cover = cover_files[0][0]

    print(f"📂 自动找到最新封面: {os.path.basename(latest_cover)}")
    return latest_cover


def get_video_properties(video_path):
    """获取视频属性（分辨率、帧率），使用 ffprobe 更轻量"""
    ffprobe = FFMPEG_PATH.replace("ffmpeg", "ffprobe")

    # 获取分辨率
    cmd_res = [
        ffprobe, "-v", "error",
        "-select_streams", "v:0",
        "-show_entries", "stream=width,height",
        "-of", "csv=s=x:p=0",
        video_path
    ]
    result = subprocess.run(cmd_res, capture_output=True, text=True)
    parts = result.stdout.strip().split('x')
    width = int(parts[0]) if len(parts) >= 2 else None
    height = int(parts[1]) if len(parts) >= 2 else None

    # 获取帧率
    cmd_fps = [
        ffprobe, "-v", "error",
        "-select_streams", "v:0",
        "-show_entries", "stream=r_frame_rate",
        "-of", "csv=s=x:p=0",
        video_path
    ]
    result = subprocess.run(cmd_fps, capture_output=True, text=True)
    fps = result.stdout.strip() or "2997/100"

    return width, height, fps


def get_media_duration(path):
    """使用 ffprobe 获取媒体时长（秒）"""
    cmd = [
        FFMPEG_PATH.replace("ffmpeg", "ffprobe"),
        "-v", "error",
        "-show_entries", "format=duration",
        "-of", "default=noprint_wrappers=1:nokey=1",
        path
    ]
    result = subprocess.run(cmd, capture_output=True, text=True, check=True)
    return float(result.stdout.strip())


def finalize_output(temp_path, output_path, expected_duration=None, tolerance=5.0):
    """校验临时输出并原子替换为正式文件"""
    if not os.path.exists(temp_path):
        raise RuntimeError(f"❌ 临时输出文件未生成: {temp_path}")

    actual_duration = get_media_duration(temp_path)
    if expected_duration and abs(actual_duration - expected_duration) > tolerance:
        raise RuntimeError(
            f"❌ 输出时长异常: {actual_duration:.2f}s，预期约 {expected_duration:.2f}s"
        )

    os.replace(temp_path, output_path)
    return actual_duration


def get_target_resolution(width, height):
    """按 OUTPUT_WIDTH 等比缩放，保持偶数高度"""
    if not width or not height:
        raise RuntimeError("❌ 无法获取视频分辨率")
    if OUTPUT_WIDTH <= 0 or width <= OUTPUT_WIDTH:
        return width, height

    target_width = OUTPUT_WIDTH
    target_height = round(height * target_width / width)
    if target_height % 2:
        target_height += 1
    return target_width, target_height


def build_video_encoding_args():
    """根据编码器生成视频编码参数。默认走 mac 硬件编码以提升速度。"""
    if VIDEO_CODEC == "h264_videotoolbox":
        return [
            "-c:v", VIDEO_CODEC,
            "-profile:v", PROFILE,
            "-level", LEVEL,
            "-b:v", VIDEO_BITRATE,
            "-realtime", "1",
            "-prio_speed", "1",
            "-pix_fmt", PIXEL_FMT,
        ]

    return [
        "-c:v", VIDEO_CODEC,
        "-profile:v", PROFILE,
        "-level", LEVEL,
        "-crf", CRF,
        "-preset", PRESET,
        "-pix_fmt", PIXEL_FMT,
    ]


def shift_srt_timestamps(srt_path, shift_ms, output_path):
    """延后 SRT 字幕时间轴

    Args:
        srt_path: 原字幕文件路径
        shift_ms: 延后毫秒数（正数延后，负数提前）
        output_path: 输出字幕文件路径
    """
    time_re = re.compile(r"(\d{2}):(\d{2}):(\d{2}),(\d{3})")

    def to_ms(h, m, s, ms):
        return ((int(h) * 60 + int(m)) * 60 + int(s)) * 1000 + int(ms)

    def from_ms(t):
        if t < 0:
            t = 0
        ms = t % 1000
        t //= 1000
        s = t % 60
        t //= 60
        m = t % 60
        t //= 60
        h = t
        return f"{h:02d}:{m:02d}:{s:02d},{ms:03d}"

    with open(srt_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    out = []
    for line in lines:
        if "-->" in line:
            parts = line.split("-->")
            if len(parts) == 2:
                a = parts[0].strip()
                b = parts[1].strip()
                ma = time_re.match(a)
                mb = time_re.match(b)
                if ma and mb:
                    ta = to_ms(*ma.groups()) + shift_ms
                    tb = to_ms(*mb.groups()) + shift_ms
                    out.append(f"{from_ms(ta)} --> {from_ms(tb)}\n")
                    continue
        out.append(line)

    with open(output_path, 'w', encoding='utf-8') as f:
        f.writelines(''.join(out))

    print(f"⏱️ 字幕时间轴延后: {shift_ms}ms")


def escape_subtitle_path(path):
    """转义字幕路径用于 ffmpeg subtitles 滤镜"""
    path = path.replace('\\', '\\\\')  # \  -> \\
    path = path.replace(':', '\\:')    # :  -> \:
    path = path.replace("'", "\\'")    # '  -> \'
    return path


def get_title_from_excel(platform_name):
    """从 Excel 读取指定平台的标题"""
    if not os.path.exists(METADATA_EXCEL):
        print(f"⚠️  元数据 Excel 不存在，使用默认命名")
        return None

    try:
        wb = load_workbook(METADATA_EXCEL)
        ws = wb.active

        # 查找对应平台的标题
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] and platform_name in str(row[0]):
                title = row[1]  # 标题在 B 列
                if title:
                    print(f"📋 从 Excel 读取标题: {title}")
                    return title

        print(f"⚠️  未在 Excel 中找到平台 [{platform_name}] 的标题")
        return None
    except Exception as e:
        print(f"⚠️  读取 Excel 失败: {e}")
        return None


def sanitize_filename(filename):
    """清理文件名，移除不合法字符"""
    # 移除或替换不合法的文件名字符
    invalid_chars = ['<', '>', ':', '"', '/', '\\', '|', '?', '*']
    for char in invalid_chars:
        filename = filename.replace(char, '')

    # 移除标题开头的日期前缀（如 "2.11"、"02.11" 等）
    import re
    # 匹配开头的数字.数字 格式
    filename = re.sub(r'^\d+\.\d+', '', filename)
    # 移除可能剩余的冒号
    filename = filename.lstrip('：:')

    return filename


def get_output_filename():
    """生成输出文件名: 3字幕+拼接后的简体标题"""
    # 从 Excel 读取油管简体标题（包含后缀）
    title = get_title_from_excel("油管简体")

    # 获取日期
    now = datetime.now()
    month_day = now.strftime("%m.%d")  # 使用 2.9 格式

    # 组合文件名
    if title:
        # 3字幕 + 日期 + 标题
        clean_title = sanitize_filename(title)
        filename = f"3字幕{month_day}{clean_title}.mp4"
        # 限制文件名长度（macOS 文件名限制 255 字符）
        if len(filename) > 200:
            filename = f"3字幕{month_day}.mp4"
        return filename
    else:
        # 备用：仅使用日期
        return f"3字幕{month_day}.mp4"


def process_video(input_video, subtitle_path, cover_path):
    """
    处理视频：1.1倍速 + 去重 + 拼接封面 + 字幕延后 + 烧录字幕（单次编码）

    处理规则：
    1. 播放速度：1.1倍（音画同步，不改变音调）
    2. 去重：Video3 专用 - 亮度 +2% (与 Video1/Video2 不同)
    3. 拼接封面：前置 0.2 秒封面
    4. 字幕时间轴：延后 0.2 秒
    5. 烧录字幕：硬字幕烧录
    6. 编码：libx264 / yuv420p / high / 4.2 / CRF 15
    7. 音频：AAC 320k

    优化：封面拼接 + 加速 + 字幕烧录在一次 ffmpeg pass 中完成，
    避免之前先编码中间文件再重新编码的问题。
    """
    output_filename = get_output_filename()
    output_path = os.path.join(OUTPUT_DIR, output_filename)
    temp_output_path = output_path + ".part.mp4"

    # 确保输出目录存在
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    if os.path.exists(temp_output_path):
        os.remove(temp_output_path)

    # 获取视频属性
    width, height, fps = get_video_properties(input_video)
    print(f"📐 视频属性: {width}x{height} @ {fps} fps")
    target_width, target_height = get_target_resolution(width, height)
    print(f"📐 输出分辨率: {target_width}x{target_height}")

    # 创建临时目录（仅用于存放延后后的字幕文件）
    temp_dir = tempfile.mkdtemp()
    try:
        temp_srt = os.path.join(temp_dir, "shifted.srt")

        # 步骤 1：延后字幕时间轴 0.2 秒（纯文本操作，极快）
        print("📝 步骤 1/2：延后字幕时间轴 0.2 秒...")
        shift_srt_timestamps(subtitle_path, 200, temp_srt)

        # 步骤 2：单次 ffmpeg pass 完成所有处理
        print("🎬 步骤 2/2：封面拼接 + 加速 + 字幕烧录（单次编码）...")

        # 转义字幕路径
        srt_escaped = escape_subtitle_path(temp_srt)
        fontdir = "/System/Library/Fonts"

        # 字幕样式
        subtitle_filter = (
            f"subtitles='{srt_escaped}':fontsdir='{fontdir}':"
            "force_style='FontName=PingFangSC-Medium,FontSize=22,"
            "PrimaryColour=&H00FFFFFF&,BackColour=&H80808080&,"
            "OutlineColour=&H00000000&,BorderStyle=1,Outline=0,Shadow=0,MarginV=30,Alignment=2'"
        )

        # 单次 filter_complex：封面缩放 → 主视频加速+裁剪+锐化 → 拼接 → 烧录字幕
        filter_complex = (
            f"[0:v]scale={target_width}:{target_height},setsar=1:1,fps={fps},format=yuv420p[vcover];"
            f"[1:v]setpts=0.909*PTS,scale={target_width}:{target_height},setsar=1:1,fps={fps},format=yuv420p[vmain];"
            f"[1:a]atempo=1.1[amain];"
            f"[vcover][vmain]concat=n=2:v=1:a=0[vconcat];"
            f"[vconcat]{subtitle_filter}[vout]"
        )

        cmd = [
            FFMPEG_PATH,
            "-loop", "1", "-t", "0.2", "-i", cover_path,
            "-i", input_video,
            "-filter_complex", filter_complex,
            "-map", "[vout]",
            "-map", "[amain]",

            # 视频编码参数
            *build_video_encoding_args(),

            # 音频编码参数
            "-c:a", AUDIO_CODEC,
            "-b:a", AUDIO_BITRATE,

            # 优化选项
            "-movflags", "+faststart",

            # 覆盖已存在的文件
            "-y",
            temp_output_path
        ]

        result = subprocess.run(cmd, capture_output=True)
        if result.returncode != 0:
            if os.path.exists(temp_output_path):
                os.remove(temp_output_path)
            error_msg = result.stderr.decode('utf-8') if result.stderr else result.stdout.decode('utf-8')
            raise RuntimeError(f"❌ ffmpeg 执行失败:\n{error_msg}")

    finally:
        # 清理临时文件
        shutil.rmtree(temp_dir, ignore_errors=True)

    expected_duration = get_media_duration(input_video) / 1.1 + 0.2
    actual_duration = finalize_output(temp_output_path, output_path, expected_duration)
    print(f"✅ 输出时长校验通过: {actual_duration:.2f}s")

    return output_path


def main():
    lock_file = None
    try:
        lock_file = acquire_process_lock()

        # 1. 校验 ffmpeg
        validate_ffmpeg()

        # 2. 解析参数或自动获取
        if len(sys.argv) >= 2:
            video_path = sys.argv[1]
            print(f"📂 使用指定视频: {video_path}")
        else:
            video_path = get_latest_video()

        if len(sys.argv) >= 3:
            subtitle_path = sys.argv[2]
            print(f"📂 使用指定字幕: {subtitle_path}")
        else:
            subtitle_path = get_latest_subtitle()

        if len(sys.argv) >= 4:
            cover_path = sys.argv[3]
            print(f"📂 使用指定封面: {cover_path}")
        else:
            cover_path = get_latest_cover()

        # 3. 字幕验证（Video3 生成前强制检查）
        print("\n" + "="*50)
        print("🔍 Video3 生成前字幕验证")
        print("="*50)
        
        try:
            from subtitle_validator import validate_subtitle
            from backup_manager import create_backup
            
            # 创建当前字幕备份
            with open(subtitle_path, 'r') as f:
                sub_content = f.read()
            sub_count = len([b for b in sub_content.split('\n\n') if '-->' in b])
            create_backup(subtitle_path, sub_count, "Video3生成前", "自动备份")
            
            # 验证字幕
            audio_path = subtitle_path.replace("/Users/ai/Documents/video_pipeline/3daily/简体", "/Users/ai/Documents/video_pipeline/3daily/audio/").replace(".srt", ".wav")
            volc_path = "/Users/ai/.claude/skills/jz字幕/volcengine_result.json"
            
            issues = validate_subtitle(subtitle_path, audio_path if os.path.exists(audio_path) else None, volc_path if os.path.exists(volc_path) else None)
            
            if issues:
                print("\n⚠️ 字幕验证发现问题，建议检查后再生成")
        except ImportError as e:
            print(f"⚠️ 无法导入验证器: {e}")
        
        # 4. 处理视频
        output_path = process_video(video_path, subtitle_path, cover_path)

        # 5. 返回结果
        print(f"\n✅ Video3 生成成功: {output_path}")
        print(f"📁 文件位置: {OUTPUT_DIR}")

        # 6. 输出标题（必须发送给用户）
        title = get_title_from_excel("油管简体")
        if title:
            print(f"\n📌 Video3 标题: {title}")

    except RuntimeError as e:
        print(str(e), file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"❌ 未知错误: {e}", file=sys.stderr)
        sys.exit(1)
    finally:
        if lock_file:
            lock_file.close()


if __name__ == "__main__":
    main()
