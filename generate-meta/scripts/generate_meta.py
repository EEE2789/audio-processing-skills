#!/usr/bin/env python3
"""
币圈行情分析文稿 - 多平台标题和简介生成工具

基于 final.txt 生成各平台的标题和简介，并写入 Excel
"""

import os
import sys
import json
import subprocess
import re
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment

# ====== 配置 ======
ENV_FILE = Path(__file__).parent.parent / ".env"
# Excel 配置文件路径（用户维护的平台配置）
CONFIG_EXCEL = Path("/Users/ai/Documents/video_pipeline/4fixed/subtitle_rules/标题和简介要求2026-02-08.xlsx")
# 输出 Excel 文件路径（最终确认的文件）
OUTPUT_DIR = Path("/Users/ai/Documents/video_pipeline/2output")
OUTPUT_FILE = OUTPUT_DIR / "视频自动上传.xlsx"
# 临时审核文件路径（放到 3daily 文件夹）
DAILY_DIR = Path("/Users/ai/Documents/video_pipeline/3daily")
REVIEW_FILE = DAILY_DIR / "元数据审核.txt"

# ====== 繁体转换（复用 jz字幕 的规则） ======
S2T_MAP = {
    '万': '萬', '亿': '億', '比特币': '比特幣', '稳定币': '穩定幣',
    '开仓': '開倉', '止损': '止損', '止盈': '止盈',
    '热力图': '熱力圖', '流动性': '流動性', '散户': '散戶', '筹码': '籌碼',
    '买': '買', '卖': '賣', '图': '圖', '线': '線', '号': '號',
    '价': '價', '现': '現', '场': '場', '势': '勢', '态': '態',
    '况': '況', '标': '標', '记': '記', '录': '錄', '汇': '匯',
    '划': '劃', '见': '見', '点': '點', '头': '頭', '长': '長',
    '门': '門', '项': '項', '题': '題', '关': '关', '类': '類',
    '种': '種', '样': '樣', '当': '當', '选': '選', '备': '備',
    '复': '復', '杂': '雜', '极': '極', '构': '構', '济': '濟',
    '营': '營', '验': '驗', '额': '額', '账': '賬', '财': '財',
    '务': '務', '损': '損', '益': '益', '赚': '賺', '赔': '賠',
    '贷': '貸', '款': '款', '准': '準', '确': '確', '误': '誤',
    '错': '錯', '币': '幣', '缠论': '纏論', '背驰': '背馳',
    '笔': '筆', '线段': '線段', '中枢': '中樞', '走势': '走勢',
    '盘整': '盤整', '趋势': '趨勢', '上涨': '上漲', '下跌': '下跌',
    '震荡': '震盪', '双底': '雙底', '双顶': '雙頂', '颈线': '頸線',
    '突破': '突破', '下破': '下破', '上破': '上破', '做多': '做多',
    '做空': '做空', '多单': '多單', '空单': '空單', '平仓': '平倉',
    '减仓': '減倉', '加仓': '加倉', '杠杆': '槓桿', '合约': '合約',
    '现货': '現貨', '期货': '期貨', '期权': '期權', '多头': '多頭',
    '空头': '空頭', '反弹': '反彈', '回调': '回調', '回踩': '回踩',
    '探底': '探底', '筑底': '築底', '拉升': '拉升', '跳水': '跳水',
    '暴涨': '暴漲', '暴跌': '暴跌', '横盘': '橫盤', '整理': '整理',
    '吸筹': '吸籌', '出货': '出貨', '诱多': '誘多', '诱空': '誘空',
    '均线': '均線', '布林带': '布林帶', '成交量': '成交量', '背离': '背離',
    '金叉': '金叉', '死叉': '死叉', '压力位': '壓力位', '支撑位': '支撐位',
    '阻力位': '阻力位', '日线': '日線', '周线': '周線', '月线': '月線',
    '小时': '小時', '分钟': '分鐘', 'K线': 'K線', '阳线': '陽線',
    '阴线': '陰线', '十字星': '十字星', '仓位': '倉位', '重仓': '重倉',
    '轻仓': '輕倉', '满仓': '滿倉', '空仓': '空倉', '风控': '風控',
    '抄底': '抄底', '逃顶': '逃頂', '追涨': '追漲', '杀跌': '殺跌',
    '后': '後', '来': '來', '去': '去', '里': '裡', '国': '國',
    '营业': '營業', '业务': '業務', '义': '義', '务': '務',
    '说明': '說明', '这个': '這個', '那个': '那個', '开始': '開始',
    '结束': '結束', '时间': '時間', '钟': '鐘',
}

def to_traditional(text):
    """简体转繁体"""
    result = text
    for s, t in S2T_MAP.items():
        result = result.replace(s, t)
    return result


# ====== 工具函数 ======

def get_current_date_short():
    """获取当前日期，格式如 2.8（月.日）"""
    now = datetime.now()
    return f"{now.month}.{now.day}"


def remove_date_prefix(text):
    """去除文本开头的日期格式（如 2.8、2.8深度分析：、2月8日深度解析：等）

    Args:
        text: 原文本

    Returns:
        去除日期前缀后的文本
    """
    # 匹配日期格式：数字.数字 开头的各种变体
    patterns = [
        r'^\d+\.\d+[：:]\s*',           # 2.8： 或 2.8:
        r'^\d+\.\d+\s*深度分析[：:]\s*',  # 2.8深度分析： 或 2.8深度分析:
        r'^\d+\.\d+\s*行情分析[：:]\s*',  # 2.8行情分析：
        r'^\d+\.\d+\s*行情[：:]\s*',     # 2.8行情：
        r'^\d+\.\d+\s*',                # 2.8 开头（带空格）
        # 匹配 "2月8日" 或 "2月8深度" 等格式
        r'^\d+月\d+日\s*',              # 2月8日
        r'^\d+月\d+日深度分析[：:]\s*',  # 2月8日深度分析：
        r'^\d+月\d+日深度解析[：:]\s*',  # 2月8日深度解析：
        r'^\d+月\d+深度[：:]\s*',        # 2月8深度：
        # 匹配剩余的 "深度分析：" 或 "深度解析："（日期已去除后残留的）
        r'^深度分析[：:]\s*',
        r'^深度解析[：:]\s*',
    ]

    result = text
    for pattern in patterns:
        result = re.sub(pattern, '', result, count=1)

    return result


def parse_limit(limit_str):
    """解析字数限制字符串，提取数字

    例如: "小于50" -> 50, "小于100" -> 100
    """
    if not limit_str:
        return 999999  # 无限制
    match = re.search(r'\d+', str(limit_str))
    if match:
        return int(match.group())
    return 999999


def load_platform_config_from_excel():
    """从 Excel 文件加载平台配置

    返回平台列表，按 Excel 中的顺序排列
    """
    if not CONFIG_EXCEL.exists():
        raise RuntimeError(f"❌ 配置文件不存在: {CONFIG_EXCEL}")

    wb = load_workbook(CONFIG_EXCEL)
    ws = wb.active

    platforms = []

    # 跳过表头，从第二行开始读取
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # 跳过空行
            continue

        # 解析行数据
        # 列: 0=渠道, 1=标题字数限制, 2=简介字数限制, 3=标题前缀, 4=标题后缀, 5=简介前缀, 6=简介后缀
        platform = {
            'name': row[0],
            'title_max': parse_limit(row[1]),
            'desc_max': parse_limit(row[2]),
            'title_prefix': row[3] or '',
            'title_suffix': row[4] or '',
            'desc_prefix': row[5] or '',
            'desc_suffix': row[6] or '',
        }

        # 检测是否需要繁体转换（根据平台名称或内容）
        if '繁体' in platform['name'] or '繁體' in platform['name']:
            platform['convert_to_traditional'] = True
        else:
            platform['convert_to_traditional'] = False

        platforms.append(platform)

    return platforms


def write_review_file(title_short, title_medium, title_long, base_desc):
    """写入审核文件"""
    DAILY_DIR.mkdir(parents=True, exist_ok=True)

    content = f"""# DeepSeek 生成内容审核

## 超短标题 ({len(title_short)}字)
{title_short}

## 中等标题 ({len(title_medium)}字)
{title_medium}

## 长标题 ({len(title_long)}字)
{title_long}

## 简介 ({len(base_desc)}字)
{base_desc}
"""

    with open(REVIEW_FILE, 'w', encoding='utf-8') as f:
        f.write(content)

    return REVIEW_FILE


def open_review_file():
    """打开审核文件（macOS）"""
    try:
        subprocess.run(['open', str(REVIEW_FILE)], check=True)
    except:
        print(f"💡 请手动打开审核文件: {REVIEW_FILE}")


def read_review_file():
    """从审核文件读取修改后的内容"""
    with open(REVIEW_FILE, 'r', encoding='utf-8') as f:
        content = f.read()

    # 解析内容
    title_short = ""
    title_medium = ""
    title_long = ""
    base_desc = ""

    current_section = None
    current_content = []

    for line in content.split('\n'):
        if line.startswith('## 超短标题'):
            if current_section == "desc":
                base_desc = '\n'.join(current_content).strip()
            current_section = "short"
            current_content = []
        elif line.startswith('## 中等标题'):
            if current_section == "short":
                title_short = '\n'.join(current_content).strip()
            current_section = "medium"
            current_content = []
        elif line.startswith('## 长标题'):
            if current_section == "medium":
                title_medium = '\n'.join(current_content).strip()
            current_section = "long"
            current_content = []
        elif line.startswith('## 简介'):
            if current_section == "long":
                title_long = '\n'.join(current_content).strip()
            current_section = "desc"
            current_content = []
        elif line.startswith('---') or line.startswith('# '):
            continue
        else:
            current_content.append(line)

    # 保存最后一个部分
    if current_section == "short":
        title_short = '\n'.join(current_content).strip()
    elif current_section == "medium":
        title_medium = '\n'.join(current_content).strip()
    elif current_section == "long":
        title_long = '\n'.join(current_content).strip()
    elif current_section == "desc":
        base_desc = '\n'.join(current_content).strip()

    return title_short, title_medium, title_long, base_desc


def load_env():
    """加载环境变量"""
    if not ENV_FILE.exists():
        return {}

    env_vars = {}
    with open(ENV_FILE, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith('#') and '=' in line:
                key, value = line.split('=', 1)
                env_vars[key.strip()] = value.strip()
    return env_vars


def validate_and_truncate(text, max_length, field_name="内容"):
    """验证文本长度，如果超过限制则自动截断

    Args:
        text: 原文本
        max_length: 最大长度
        field_name: 字段名称（用于日志）

    Returns:
        原文本（如果符合限制）或截断后的文本
    """
    if len(text) <= max_length:
        return text

    # 超过限制，自动截断
    truncated = text[:max_length]
    print(f"  ⚠️ {field_name}超长，已截断：{len(text)}字 → {max_length}字")
    return truncated


def call_deepseek(content):
    """调用 DeepSeek 生成标题和简介"""
    env_vars = load_env()
    api_key = env_vars.get('DEEPSEEK_API_KEY')

    if not api_key:
        raise RuntimeError("❌ 未找到 DEEPSEEK_API_KEY，请在 .env 文件中设置")

    # ============ DeepSeek Prompt (用户可优化) ============
    prompt = f'''你现在是一个币圈比特币和以太坊的行情分析师，同时具备优秀的文案编辑能力。我待会发你一篇文章，请根据我发你的文章，请帮我用第一人称的方式，但不要出现我或我们等字眼，写一份总字数在300字到400字之间的内容要点，每小段在80字以上，150字以内，请分段展示，不要使用列表。
你根据我发你的文章，总结的内容请按以下结果展现：

比特币行情分析：主要是波浪结构的分析。不需要自我介绍部分，以下文字不要出现以下内容："大家好，我是军长"
如果文档中出现较长的以太坊内容，增加以太坊分析，如果没有出现，就不写。
开单建议

请一定要按照我发你的文章总结。文章的受众对象是币圈的合约玩家，请按照段落的形式展现，不要列表。文章内容如下：


## 输入文稿
{content}

## 输出要求
请严格按以下 JSON 格式输出（不要有其他内容）：
{{
  "title_short": "（3-8字的极短标题，用于微博等限制30字以下的平台。**重要**：标题会自动添加前后缀约22字，因此标题内容请严格控制在3-8字，总长不超过30字。示例："B浪反弹中"、"等回调做多"、"耐心等待"）",
  "title_medium": "（20-30字的中等标题，用于知乎/推特等限制50-80字的平台。**重要**：标题会自动添加前后缀约21字，因此标题内容请控制在20-30字）",
  "title_long": "（45-55字的长标题，用于油管/B站/Facebook等限制80字以上的平台。**重要**：标题会自动添加日期前缀和后缀，预留约30-35字，因此标题内容请控制在45-55字）",
  "description": "（250-300字的简介，包含分析要点和风险提示。**重要**：简介必须分段显示，每段80-150字，段落之间用双换行符\\n\\n分隔（段落间空一行），方便手机阅读）**简介结尾不要包含**："说明：直接修改上述内容，保存后退出即可。" 这类指导性文字"
}}

## 内容要求
1. 基于文稿真实提炼，不夸大、不编造
2. 保持理性分析风格
3. 标题不要以日期开头（如"2.8行情"、"2.8深度分析"等），这些会自动添加
4. 标题不要以"行情"、"深度分析"、"深度解析"等开头
5. 标题中不要使用冒号（：），不要用"XX分析："、"XX解读："等格式
6. **标题使用正常标点符号**：句子内用逗号（，）分隔分句，句子结尾不要用句号等标点（让标题感觉更自然）
7. **禁止使用时间级别描述**：不要在标题和简介中使用"2小时级别"、"4小时级别"、"日线级别"等时间级别表述
8. 简介必须分段，每段80-150字，段落间用\\n\\n分隔（段落间空一行），方便手机阅读
9. 简介结尾必须加风险提示
10. **严格遵守字数限制**：
    - 超短标题：5-10字（用于微博等限制30字以下的平台）
    - 中等标题：20-30字（用于知乎/推特等限制50-80字的平台）
    - 长标题：45-55字（用于油管/B站/Facebook等限制80字以上的平台）
    - 简介：250-300字（不含前后缀）
11. **禁止使用省略号或截断**：标题必须完整，不要用"..."、"等等"、"等"来截断内容
12. **币种名称完整**：不要缩写币种名称（如"ne"应写为"NEAR"、"XLM"应完整写出）
13. **精简表达**：删除冗余词汇，如"目前"、"当前"、"现在"、"一直"、"始终"等可省略的词语
14. **超短标题要极简**：微博的标题内容只有5-10字空间，必须极度精简，如"比特币B浪反弹，等回调"
15. **禁止推广和引导类内容**：
    - 绝对禁止提及：社区、群组、微信群、QQ群、Telegram群、Discord、加入群、入群
    - 绝对禁止提及：我们会通知大家、及时通知、关注社区、加入我们
    - 绝对禁止提及：免费开放、全部免费、免费建议、交易建议
    - 绝对禁止提及：联系方式、联系我们、私信我们
    - 如文稿中出现此类内容，请自动忽略或转换为纯技术分析表述
'''

    # 调用 DeepSeek API
    try:
        result = subprocess.run(
            ['curl', '-s', '-X', 'POST',
             'https://api.deepseek.com/v1/chat/completions',
             '-H', 'Content-Type: application/json',
             '-H', f'Authorization: Bearer {api_key}',
             '-d', json.dumps({
                 "model": "deepseek-chat",
                 "messages": [{"role": "user", "content": prompt}],
                 "temperature": 0.7
             })],
            capture_output=True,
            text=True,
            timeout=120
        )

        if result.returncode != 0:
            raise RuntimeError(f"DeepSeek API 调用失败: {result.stderr}")

        response = json.loads(result.stdout)

        if 'choices' not in response or len(response['choices']) == 0:
            raise RuntimeError("DeepSeek API 返回格式错误")

        content = response['choices'][0]['message']['content']

        # 解析 JSON
        try:
            result = json.loads(content)
            title_short = result.get('title_short', '')
            title_medium = result.get('title_medium', result.get('title_short', ''))
            title_long = result.get('title_long', result.get('title', ''))
            description = result.get('description', '')
            return title_short, title_medium, title_long, description
        except json.JSONDecodeError:
            # 如果返回的不是纯 JSON，尝试提取 JSON 部分
            import re
            json_match = re.search(r'\{[^{}]*\}', content)
            if json_match:
                result = json.loads(json_match.group())
                title_short = result.get('title_short', '')
                title_medium = result.get('title_medium', result.get('title_short', ''))
                title_long = result.get('title_long', result.get('title', ''))
                description = result.get('description', '')
                return title_short, title_medium, title_long, description
            else:
                raise RuntimeError(f"无法解析 DeepSeek 返回的内容: {content}")

    except subprocess.TimeoutExpired:
        raise RuntimeError("DeepSeek API 调用超时")
    except Exception as e:
        raise RuntimeError(f"调用 DeepSeek 失败: {e}")


def process_platform(title_short, title_medium, title_long, base_desc, platform_config, current_date):
    """处理单个平台的内容

    Args:
        title_short: 超短标题（10-15字）
        title_medium: 中等标题（20-30字）
        title_long: 长标题（45-55字）
        base_desc: 基础简介
        platform_config: 平台配置
        current_date: 当前日期（格式：2.8）
    """

    # 获取平台名称（用于错误提示）
    platform_name = platform_config.get('name', '')

    # 根据平台字数限制选择标题
    title_limit = platform_config['title_max']
    if title_limit <= 30:
        # 微博（30字）：使用超短标题
        base_title = title_short
    elif title_limit <= 50:
        # 知乎（50字）：使用中等标题
        base_title = title_medium
    else:
        # 油管/B站/Facebook（80-100字）：使用长标题
        base_title = title_long

    # 去除 DeepSeek 生成标题中的日期前缀
    base_title = remove_date_prefix(base_title)

    # 处理前缀：将 "日期+" 替换为实际日期
    title_prefix = platform_config.get('title_prefix', '')
    if '日期+' in title_prefix:
        title_prefix = title_prefix.replace('日期+', current_date)

    # 组装标题：前缀 + 基础标题 + 后缀
    title = title_prefix + base_title + platform_config.get('title_suffix', '')
    title = validate_and_truncate(title, title_limit, f"{platform_name}标题")

    # 处理简介前缀：将 "日期+" 替换为实际日期
    desc_prefix = platform_config.get('desc_prefix', '')
    if '日期+' in desc_prefix:
        desc_prefix = desc_prefix.replace('日期+', current_date)
    desc_limit = platform_config['desc_max']

    # Twitter 特殊处理：如果 base_desc 开头已有 $BTC，则移除（避免与前缀重复）
    if platform_name == '推特' or platform_name == 'twitter' or platform_name == 'Twitter':
        base_desc = base_desc.strip()
        if base_desc.startswith('$BTC ') or base_desc.startswith('$BTC\n'):
            base_desc = base_desc[5:].lstrip()  # 移除开头的 "$BTC " 或 "$BTC\n"
        elif base_desc.startswith('$BTC'):
            base_desc = base_desc[4:].lstrip()  # 移除开头的 "$BTC"

    # 组装简介：前缀 + 基础简介 + 后缀
    desc = desc_prefix + base_desc + platform_config.get('desc_suffix', '')
    desc = validate_and_truncate(desc, desc_limit, f"{platform_name}简介")

    # 繁体转换
    if platform_config.get('convert_to_traditional', False):
        title = to_traditional(title)
        desc = to_traditional(desc)

    return title, desc


def write_excel(platforms_data, platforms_config):
    """写入 Excel 文件

    Args:
        platforms_data: [(platform_name, title, desc), ...]
        platforms_config: 从 Excel 读取的平台配置列表
    """
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 写入表头
    headers = ['渠道', '标题', '简介', '字数', '系统限制', '标题', '标题限制']
    ws.append(headers)

    # 写入数据（按照 platforms_config 的顺序）
    for idx, (platform_data, platform_config) in enumerate(zip(platforms_data, platforms_config), start=2):
        platform_name, title, desc = platform_data

        row_data = [
            platform_name,
            title,
            desc,
            f'=LEN(C{idx})',
            f'小于{platform_config["desc_max"]}',
            f'=LEN(B{idx})',
            f'小于{platform_config["title_max"]}'
        ]
        ws.append(row_data)

    # 保存文件
    wb.save(OUTPUT_FILE)
    return OUTPUT_FILE


# ====== 主逻辑 ======

def main():
    # 解析参数
    # 默认自动模式，不需要 --auto 参数
    auto_mode = True

    # 标记是否已从审核文件读取
    use_review_content = False

    # 如果是自动模式，先检查是否有审核文件
    if auto_mode:
        review_file_path = Path(REVIEW_FILE)
        if review_file_path.exists():
            print(f"📖 发现审核文件，读取修改后的内容...")
            print(f"📂 文件路径: {REVIEW_FILE}")
            # 从审核文件读取
            title_short, title_long, base_desc = read_review_file()
            use_review_content = True
            print(f"\n✓ 超短标题 ({len(title_short)}字): {title_short}")
            print(f"✓ 中等标题 ({len(title_medium)}字): {title_medium}")
            print(f"✓ 长标题 ({len(title_long)}字): {title_long}")
            print(f"✓ 简介 ({len(base_desc)}字): ", end="")
            print(f"{base_desc[:100]}..." if len(base_desc) > 100 else base_desc)
            print("\n" + "="*60)
            print("⚡ 使用审核文件中的内容继续生成...")
            print("="*60)

    # 获取当前日期
    current_date = get_current_date_short()
    print(f"📅 使用日期: {current_date}")

    # 自动读取 final.txt
    default_txt_path = Path("/Users/ai/Documents/video_pipeline/3daily/final.txt")

    # 如果命令行指定了路径，使用指定的；否则使用默认路径
    txt_file = None
    for arg in sys.argv[1:]:
        if not arg.startswith('--') and not arg.startswith('-'):
            txt_file = Path(arg)
            break

    if txt_file is None:
        txt_file = default_txt_path

    if not txt_file.exists():
        print(f"❌ 文件不存在: {txt_file}")
        print(f"💡 提示：可以将 final.txt 放到 {default_txt_path}")
        sys.exit(1)

    # 读取文稿
    print(f"📄 读取文稿: {txt_file}")
    with open(txt_file, 'r', encoding='utf-8') as f:
        content = f.read()

    if len(content.strip()) < 50:
        print("❌ 文稿内容太短，请检查文件")
        sys.exit(1)

    # 加载平台配置（从 Excel）
    print(f"📋 读取平台配置: {CONFIG_EXCEL}")
    try:
        platforms_config = load_platform_config_from_excel()
        print(f"  ✓ 已加载 {len(platforms_config)} 个平台配置")
        for p in platforms_config:
            print(f"    - {p['name']}: 标题≤{p['title_max']}字, 简介≤{p['desc_max']}字")
    except RuntimeError as e:
        print(f"\n{e}")
        sys.exit(1)

    # 如果没有从审核文件读取，则调用 DeepSeek 生成
    if not use_review_content:
        # 调用 DeepSeek 生成
        print("\n🤖 正在调用 DeepSeek 生成标题和简介...")
        try:
            title_short, title_medium, title_long, base_desc = call_deepseek(content)
        except RuntimeError as e:
            print(f"\n{e}")
            sys.exit(1)

    # 显示预览
    print("\n" + "="*60)
    print("📝 生成结果预览")
    print("="*60)
    print(f"\n超短标题 ({len(title_short)}字):")
    print(f"  {title_short}")
    print(f"\n中等标题 ({len(title_medium)}字):")
    print(f"  {title_medium}")
    print(f"\n长标题 ({len(title_long)}字):")
    print(f"  {title_long}")
    print(f"\n简介 ({len(base_desc)}字):")
    print(f"  {base_desc[:100]}..." if len(base_desc) > 100 else f"  {base_desc}")
    print("\n" + "="*60)
    print("💡 说明：微博使用超短标题，知乎/推特使用中等标题，油管/B站/Facebook使用长标题")
    print("💡 前后缀将在下一步添加")
    print("\n⚡ 自动模式：跳过确认，直接生成...")

    # 处理所有平台
    print("\n🔄 正在处理各平台内容...")
    platforms_data = []
    for platform_config in platforms_config:
        platform_name = platform_config['name']
        title, desc = process_platform(title_short, title_medium, title_long, base_desc, platform_config, current_date)
        platforms_data.append((platform_name, title, desc))
        print(f"  ✓ {platform_name}: 标题{len(title)}字, 简介{len(desc)}字")

    # 用户审核通过后直接写入，无需再次确认
    print("\n✅ 审核通过，准备写入 Excel...")

    # 写入 Excel
    print("\n📊 正在写入 Excel...")
    output_file = write_excel(platforms_data, platforms_config)

    print(f"\n✅ 完成！")
    print(f"📁 输出文件: {output_file}")


if __name__ == "__main__":
    main()
