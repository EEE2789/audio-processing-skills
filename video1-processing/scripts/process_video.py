#!/usr/bin/env python3
"""
Video1 油管繁体｜行情视频基础处理（1.1x 去重 标准版）

对原始行情视频进行基础处理，生成 Video1 成品视频。
作为后续 Video2 / Video3 的统一画面基础。

输入：/Users/ai/Documents/video_pipeline/1input/ 中的视频文件
输出：/Users/ai/Documents/video_pipeline/2output/1繁体MMDD.mp4
"""

import sys
import os
import subprocess
import fcntl
from datetime import datetime
from openpyxl import load_workbook

# ====== 固定配置 ======
FFMPEG_PATH = "/opt/homebrew/bin/ffmpeg"
INPUT_DIR = "/Users/ai/Documents/video_pipeline/1input"
OUTPUT_DIR = "/Users/ai/Documents/video_pipeline/2output"
METADATA_EXCEL = "/Users/ai/Documents/video_pipeline/2output/视频自动上传.xlsx"

# 编码参数（固定）
VIDEO_CODEC = os.getenv("VIDEO_CODEC", "h264_videotoolbox")
VIDEO_BITRATE = os.getenv("VIDEO_BITRATE", "8000k")
OUTPUT_WIDTH = int(os.getenv("OUTPUT_WIDTH", "1920"))
PIXEL_FMT = "yuv420p"
PROFILE = "high"
LEVEL = "5.2"
CRF = "15"
PRESET = "veryfast"
FPS = "24"
AUDIO_CODEC = "aac"
AUDIO_BITRATE = "320k"


def acquire_process_lock():
    """防止同一个 Video1 任务被并发启动"""
    lock_path = "/tmp/video_pipeline_video1.lock"
    lock_file = open(lock_path, "w")
    try:
        fcntl.flock(lock_file, fcntl.LOCK_EX | fcntl.LOCK_NB)
    except BlockingIOError:
        raise RuntimeError("❌ Video1 正在生成中，请等待当前任务完成后再运行")
    return lock_file


def validate_ffmpeg():
    """校验 ffmpeg 是否存在"""
    if not os.path.exists(FFMPEG_PATH):
        raise RuntimeError(f"❌ ffmpeg 不存在于指定路径: {FFMPEG_PATH}")


def get_latest_video():
    """自动获取 1input 文件夹中最新的视频文件"""
    if not os.path.exists(INPUT_DIR):
        raise RuntimeError(f"❌ 输入目录不存在: {INPUT_DIR}")

    # 支持的视频格式
    video_extensions = ['.mp4', '.mov', '.avi', '.mkv', '.flv', '.wmv', '.m4v']

    # 获取所有视频文件
    video_files = []
    for file in os.listdir(INPUT_DIR):
        if any(file.lower().endswith(ext) for ext in video_extensions):
            file_path = os.path.join(INPUT_DIR, file)
            video_files.append((file_path, os.path.getmtime(file_path)))

    if not video_files:
        raise RuntimeError(f"❌ 在 {INPUT_DIR} 中未找到视频文件")

    # 按修改时间排序，获取最新的
    video_files.sort(key=lambda x: x[1], reverse=True)
    latest_video = video_files[0][0]

    print(f"📂 自动找到最新视频: {os.path.basename(latest_video)}")
    return latest_video


def validate_video(video_path):
    """校验视频文件"""
    if not os.path.exists(video_path):
        raise RuntimeError(f"❌ 视频文件不存在: {video_path}")

    # 检查是否有视频流
    cmd = [
        FFMPEG_PATH,
        "-i", video_path,
        "-hide_banner"
    ]
    result = subprocess.run(cmd, capture_output=True)
    output = result.stdout.decode('utf-8') + result.stderr.decode('utf-8')

    if "Video:" not in output:
        raise RuntimeError(f"❌ 视频文件无效或无视频流: {video_path}")

    return True


def get_media_duration(path):
    """使用 ffprobe 获取媒体时长（秒）"""
    cmd = [
        FFMPEG_PATH.replace("ffmpeg", "ffprobe"),
        "-v", "error",
        "-show_entries", "format=duration",
        "-of", "default=noprint_wrappers=1:nokey=1",
        path
    ]
    result = subprocess.run(cmd, capture_output=True, text=True, check=True)
    return float(result.stdout.strip())


def get_video_dimensions(path):
    """获取视频分辨率"""
    cmd = [
        FFMPEG_PATH.replace("ffmpeg", "ffprobe"),
        "-v", "error",
        "-select_streams", "v:0",
        "-show_entries", "stream=width,height",
        "-of", "csv=s=x:p=0",
        path
    ]
    result = subprocess.run(cmd, capture_output=True, text=True, check=True)
    width, height = result.stdout.strip().split("x")
    return int(width), int(height)


def get_target_dimensions(path):
    """按 OUTPUT_WIDTH 等比缩放，保持偶数高度"""
    width, height = get_video_dimensions(path)
    if OUTPUT_WIDTH <= 0 or width <= OUTPUT_WIDTH:
        return width, height

    target_width = OUTPUT_WIDTH
    target_height = round(height * target_width / width)
    if target_height % 2:
        target_height += 1
    return target_width, target_height


def finalize_output(temp_path, output_path, expected_duration=None, tolerance=5.0):
    """校验临时输出并原子替换为正式文件"""
    if not os.path.exists(temp_path):
        raise RuntimeError(f"❌ 临时输出文件未生成: {temp_path}")

    actual_duration = get_media_duration(temp_path)
    if expected_duration and abs(actual_duration - expected_duration) > tolerance:
        raise RuntimeError(
            f"❌ 输出时长异常: {actual_duration:.2f}s，预期约 {expected_duration:.2f}s"
        )

    os.replace(temp_path, output_path)
    return actual_duration


def build_video_encoding_args():
    """根据编码器生成视频编码参数。默认走 mac 硬件编码以提升速度。"""
    if VIDEO_CODEC == "h264_videotoolbox":
        return [
            "-c:v", VIDEO_CODEC,
            "-profile:v", PROFILE,
            "-level", LEVEL,
            "-b:v", VIDEO_BITRATE,
            "-realtime", "1",
            "-prio_speed", "1",
            "-pix_fmt", PIXEL_FMT,
        ]

    return [
        "-c:v", VIDEO_CODEC,
        "-profile:v", PROFILE,
        "-level", LEVEL,
        "-crf", CRF,
        "-preset", PRESET,
        "-pix_fmt", PIXEL_FMT,
    ]


def get_title_from_excel(platform_name):
    """从 Excel 读取指定平台的标题"""
    if not os.path.exists(METADATA_EXCEL):
        print(f"⚠️  元数据 Excel 不存在，使用默认命名")
        return None

    try:
        wb = load_workbook(METADATA_EXCEL)
        ws = wb.active

        # 查找对应平台的标题
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] and platform_name in str(row[0]):
                title = row[1]  # 标题在 B 列
                if title:
                    print(f"📋 从 Excel 读取标题: {title}")
                    return title

        print(f"⚠️  未在 Excel 中找到平台 [{platform_name}] 的标题")
        return None
    except Exception as e:
        print(f"⚠️  读取 Excel 失败: {e}")
        return None


def sanitize_filename(filename):
    """清理文件名，移除不合法字符"""
    # 移除或替换不合法的文件名字符
    invalid_chars = ['<', '>', ':', '"', '/', '\\', '|', '?', '*']
    for char in invalid_chars:
        filename = filename.replace(char, '')

    # 移除标题开头的日期前缀（如 "2.11"、"02.11"、"2.11比特幣價格走勢分析：" 等）
    import re
    # 匹配开头的数字.数字 格式
    filename = re.sub(r'^\d+\.\d+', '', filename)
    # 移除可能剩余的冒号
    filename = filename.lstrip('：:')

    return filename


def get_output_filename():
    """生成输出文件名: 1繁体+拼接后的繁体标题"""
    # 从 Excel 读取油管繁体标题
    title = get_title_from_excel("油管繁体")

    # 获取日期
    now = datetime.now()
    month_day = now.strftime("%m.%d")  # 使用 2.9 格式

    # 组合文件名
    if title:
        # 1繁体 + 日期 + 标题
        clean_title = sanitize_filename(title)
        filename = f"1繁体{month_day}{clean_title}.mp4"
        # 限制文件名长度（macOS 文件名限制 255 字符）
        if len(filename) > 200:
            filename = f"1繁体{month_day}.mp4"
        return filename
    else:
        # 备用：仅使用日期
        return f"1繁体{month_day}.mp4"


def process_video(input_video):
    """
    处理视频：1.1倍速 + 轻度去重 + 标准编码

    处理规则：
    1. 播放速度：1.1倍（音画同步，不改变音调）
    2. 帧率：29.97 fps
    3. 去重：轻微亮度扰动（使用 hue 滤镜）
    4. 编码：libx264 / yuv420p / high / 4.2 / CRF 15
    5. 音频：AAC 192k
    """
    output_filename = get_output_filename()
    output_path = os.path.join(OUTPUT_DIR, output_filename)
    temp_output_path = output_path + ".part.mp4"

    # 确保输出目录存在
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    if os.path.exists(temp_output_path):
        os.remove(temp_output_path)

    # 构建 ffmpeg 命令
    # filter_complex 说明：
    # - [0:v]setpts=0.909*PTS: 视频加速到 1.1倍 (1/1.1 ≈ 0.909)
    # - crop: 轻微裁剪 1% 画面（去重）
    # - unsharp: 轻微锐化（去重）
    # - [0:a]atempo=1.1: 音频加速到 1.1倍（保持音调）
    #
    # 去重策略：轻微裁剪 + 锐化，不影响K线颜色准确性

    target_width, target_height = get_target_dimensions(input_video)
    filter_complex = (
        f"[0:v]setpts=0.909*PTS,scale={target_width}:{target_height},setsar=1[vout];"
        "[0:a]atempo=1.1[aout]"
    )

    cmd = [
        FFMPEG_PATH,
        "-i", input_video,
        "-filter_complex", filter_complex,
        "-map", "[vout]",
        "-map", "[aout]",

        # 视频编码参数
        *build_video_encoding_args(),
        "-r", FPS,

        # 音频编码参数
        "-c:a", AUDIO_CODEC,
        "-b:a", AUDIO_BITRATE,

        # 优化选项
        "-movflags", "+faststart",

        # 覆盖已存在的文件
        "-y",
        temp_output_path
    ]

    print(f"🎬 正在处理视频...")
    print(f"📝 输入: {input_video}")
    print(f"📝 输出: {output_path}")
    print(f"⚙️  参数: 1.1倍速 | {target_width}x{target_height} | {FPS} fps | {VIDEO_CODEC} | 快速模式")

    result = subprocess.run(cmd, capture_output=True)

    if result.returncode != 0:
        if os.path.exists(temp_output_path):
            os.remove(temp_output_path)
        error_msg = result.stderr.decode('utf-8') if result.stderr else result.stdout.decode('utf-8')
        raise RuntimeError(f"❌ ffmpeg 执行失败:\n{error_msg}")

    expected_duration = get_media_duration(input_video) / 1.1
    actual_duration = finalize_output(temp_output_path, output_path, expected_duration)
    print(f"✅ 输出时长校验通过: {actual_duration:.2f}s")

    return output_path


def main():
    lock_file = None
    try:
        lock_file = acquire_process_lock()

        # 1. 校验 ffmpeg
        validate_ffmpeg()

        # 2. 自动获取最新视频（如果命令行指定了路径则使用指定路径）
        if len(sys.argv) >= 2:
            video_path = sys.argv[1]
            print(f"📂 使用指定视频: {video_path}")
        else:
            video_path = get_latest_video()

        # 3. 校验视频
        validate_video(video_path)

        # 4. 处理视频
        output_path = process_video(video_path)

        # 5. 返回结果
        print(f"\n✅ Video1 生成成功: {output_path}")
        print(f"📁 文件位置: {OUTPUT_DIR}")

    except RuntimeError as e:
        print(str(e), file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"❌ 未知错误: {e}", file=sys.stderr)
        sys.exit(1)
    finally:
        if lock_file:
            lock_file.close()


if __name__ == "__main__":
    main()
