#!/usr/bin/env python3
"""
Video2 简体｜行情视频简体版（1.1x 去重 + 结尾提示）

基于原始视频生成 Video2 成品视频。
包含结尾拼接与文字提示，用于简体渠道分发。

输入：/Users/ai/Documents/video_pipeline/1input/ 中的视频文件
输出：/Users/ai/Documents/video_pipeline/2output/2简体MMDD.mp4
"""

import sys
import os
import subprocess
import fcntl
from datetime import datetime
from openpyxl import load_workbook

# ====== 固定配置 ======
FFMPEG_PATH = "/opt/homebrew/bin/ffmpeg"
INPUT_DIR = "/Users/ai/Documents/video_pipeline/1input"
OUTPUT_DIR = "/Users/ai/Documents/video_pipeline/2output"
OUTRO_VIDEO = "/Users/ai/Documents/video_pipeline/4fixed/assets/outro_fix.mp4"
METADATA_EXCEL = "/Users/ai/Documents/video_pipeline/2output/视频自动上传.xlsx"

# 编码参数（固定）
VIDEO_CODEC = os.getenv("VIDEO_CODEC", "h264_videotoolbox")
VIDEO_BITRATE = os.getenv("VIDEO_BITRATE", "8000k")
OUTPUT_WIDTH = int(os.getenv("OUTPUT_WIDTH", "1920"))
PIXEL_FMT = "yuv420p"
PROFILE = "high"
LEVEL = "5.2"
CRF = "15"
PRESET = "veryfast"
FPS = "24"
AUDIO_CODEC = "aac"
AUDIO_BITRATE = "320k"


def acquire_process_lock():
    """防止同一个 Video2 任务被并发启动"""
    lock_path = "/tmp/video_pipeline_video2.lock"
    lock_file = open(lock_path, "w")
    try:
        fcntl.flock(lock_file, fcntl.LOCK_EX | fcntl.LOCK_NB)
    except BlockingIOError:
        raise RuntimeError("❌ Video2 正在生成中，请等待当前任务完成后再运行")
    return lock_file


def validate_ffmpeg():
    """校验 ffmpeg 是否存在"""
    if not os.path.exists(FFMPEG_PATH):
        raise RuntimeError(f"❌ ffmpeg 不存在于指定路径: {FFMPEG_PATH}")


def get_latest_video():
    """自动获取 1input 文件夹中最新的视频文件"""
    if not os.path.exists(INPUT_DIR):
        raise RuntimeError(f"❌ 输入目录不存在: {INPUT_DIR}")

    # 支持的视频格式
    video_extensions = ['.mp4', '.mov', '.avi', '.mkv', '.flv', '.wmv', '.m4v']

    # 获取所有视频文件
    video_files = []
    for file in os.listdir(INPUT_DIR):
        if any(file.lower().endswith(ext) for ext in video_extensions):
            file_path = os.path.join(INPUT_DIR, file)
            video_files.append((file_path, os.path.getmtime(file_path)))

    if not video_files:
        raise RuntimeError(f"❌ 在 {INPUT_DIR} 中未找到视频文件")

    # 按修改时间排序，获取最新的
    video_files.sort(key=lambda x: x[1], reverse=True)
    latest_video = video_files[0][0]

    print(f"📂 自动找到最新视频: {os.path.basename(latest_video)}")
    return latest_video


def validate_video(video_path):
    """校验视频文件"""
    if not os.path.exists(video_path):
        raise RuntimeError(f"❌ 视频文件不存在: {video_path}")

    # 检查是否有视频流
    cmd = [
        FFMPEG_PATH,
        "-i", video_path,
        "-hide_banner"
    ]
    result = subprocess.run(cmd, capture_output=True)
    output = result.stdout.decode('utf-8') + result.stderr.decode('utf-8')

    if "Video:" not in output:
        raise RuntimeError(f"❌ 视频文件无效或无视频流: {video_path}")

    # 检查结尾视频是否存在
    if not os.path.exists(OUTRO_VIDEO):
        raise RuntimeError(f"❌ 结尾视频不存在: {OUTRO_VIDEO}")

    return True


def get_media_duration(path):
    """使用 ffprobe 获取媒体时长（秒）"""
    cmd = [
        FFMPEG_PATH.replace("ffmpeg", "ffprobe"),
        "-v", "error",
        "-show_entries", "format=duration",
        "-of", "default=noprint_wrappers=1:nokey=1",
        path
    ]
    result = subprocess.run(cmd, capture_output=True, text=True, check=True)
    return float(result.stdout.strip())


def finalize_output(temp_path, output_path, expected_duration=None, tolerance=5.0):
    """校验临时输出并原子替换为正式文件"""
    if not os.path.exists(temp_path):
        raise RuntimeError(f"❌ 临时输出文件未生成: {temp_path}")

    actual_duration = get_media_duration(temp_path)
    if expected_duration and abs(actual_duration - expected_duration) > tolerance:
        raise RuntimeError(
            f"❌ 输出时长异常: {actual_duration:.2f}s，预期约 {expected_duration:.2f}s"
        )

    os.replace(temp_path, output_path)
    return actual_duration


def get_target_resolution(width, height):
    """按 OUTPUT_WIDTH 等比缩放，保持偶数高度"""
    if not width or not height:
        raise RuntimeError("❌ 无法获取视频分辨率")
    if OUTPUT_WIDTH <= 0 or width <= OUTPUT_WIDTH:
        return width, height

    target_width = OUTPUT_WIDTH
    target_height = round(height * target_width / width)
    if target_height % 2:
        target_height += 1
    return target_width, target_height


def build_video_encoding_args():
    """根据编码器生成视频编码参数。默认走 mac 硬件编码以提升速度。"""
    if VIDEO_CODEC == "h264_videotoolbox":
        return [
            "-c:v", VIDEO_CODEC,
            "-profile:v", PROFILE,
            "-level", LEVEL,
            "-b:v", VIDEO_BITRATE,
            "-realtime", "1",
            "-prio_speed", "1",
            "-pix_fmt", PIXEL_FMT,
        ]

    return [
        "-c:v", VIDEO_CODEC,
        "-profile:v", PROFILE,
        "-level", LEVEL,
        "-crf", CRF,
        "-preset", PRESET,
        "-pix_fmt", PIXEL_FMT,
    ]


def get_title_from_excel(platform_name):
    """从 Excel 读取指定平台的标题"""
    if not os.path.exists(METADATA_EXCEL):
        print(f"⚠️  元数据 Excel 不存在，使用默认命名")
        return None

    try:
        wb = load_workbook(METADATA_EXCEL)
        ws = wb.active

        # 查找对应平台的标题
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] and platform_name in str(row[0]):
                title = row[1]  # 标题在 B 列
                if title:
                    print(f"📋 从 Excel 读取标题: {title}")
                    return title

        print(f"⚠️  未在 Excel 中找到平台 [{platform_name}] 的标题")
        return None
    except Exception as e:
        print(f"⚠️  读取 Excel 失败: {e}")
        return None


def sanitize_filename(filename):
    """清理文件名，移除不合法字符"""
    # 移除或替换不合法的文件名字符
    invalid_chars = ['<', '>', ':', '"', '/', '\\', '|', '?', '*']
    for char in invalid_chars:
        filename = filename.replace(char, '')

    # 移除标题开头的日期前缀（如 "2.11"、"02.11" 等）
    import re
    # 匹配开头的数字.数字 格式
    filename = re.sub(r'^\d+\.\d+', '', filename)
    # 移除可能剩余的冒号
    filename = filename.lstrip('：:')

    return filename


def get_output_filename():
    """生成输出文件名: 2简体+拼接后的油管简体标题"""
    # 从 Excel 读取油管简体标题
    title = get_title_from_excel("油管简体")

    # 获取日期
    now = datetime.now()
    month_day = now.strftime("%m.%d")  # 使用 2.9 格式

    # 组合文件名
    if title:
        # 2简体 + 日期 + 标题（标题中的日期前缀会被自动移除）
        clean_title = sanitize_filename(title)
        filename = f"2简体{month_day}{clean_title}.mp4"
        # 限制文件名长度（macOS 文件名限制 255 字符）
        if len(filename) > 200:
            filename = f"2简体{month_day}.mp4"
        return filename
    else:
        # 备用：仅使用日期
        return f"2简体{month_day}.mp4"


def get_video_resolution(video_path):
    """获取视频分辨率（使用 ffprobe，比 ffmpeg -i 更轻量）"""
    cmd = [
        FFMPEG_PATH.replace("ffmpeg", "ffprobe"),
        "-v", "error",
        "-select_streams", "v:0",
        "-show_entries", "stream=width,height",
        "-of", "csv=s=x:p=0",
        video_path
    ]
    result = subprocess.run(cmd, capture_output=True, text=True)
    output = result.stdout.strip()
    if 'x' in output:
        parts = output.split('x')
        return int(parts[0]), int(parts[1])
    return None, None


def process_video(input_video):
    """
    处理视频：1.1倍速 + 轻度去重 + 文字提示 + 结尾拼接 + 标准编码

    处理规则：
    1. 播放速度：1.1倍（音画同步，不改变音调）
    2. 帧率：29.97 fps
    3. 去重：Video2 专用 - 对比度 +2% (与 Video1/Video3 不同)
    4. 文字提示：第20秒左右显示「结尾有惊喜」3秒
    5. 结尾拼接：拼接 outro_fix.mp4
    6. 编码：libx264 / yuv420p / high / 4.2 / CRF 15
    7. 音频：AAC 192k

    去重策略差异化（避免与 Video1/Video3 重复）：
    - Video1: 饱和度 +2% (eq=saturation=1.02)
    - Video2: 对比度 +2% (eq=contrast=1.02)
    - Video3: 亮度 +2% (eq=brightness=0.02)
    """
    output_filename = get_output_filename()
    output_path = os.path.join(OUTPUT_DIR, output_filename)
    temp_output_path = output_path + ".part.mp4"

    # 确保输出目录存在
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    if os.path.exists(temp_output_path):
        os.remove(temp_output_path)

    # 获取原视频分辨率
    orig_width, orig_height = get_video_resolution(input_video)
    print(f"📐 原视频分辨率: {orig_width}x{orig_height}")
    target_width, target_height = get_target_resolution(orig_width, orig_height)
    print(f"📐 输出分辨率: {target_width}x{target_height}")

    # 获取结尾视频分辨率
    outro_width, outro_height = get_video_resolution(OUTRO_VIDEO)
    print(f"📐 结尾视频分辨率: {outro_width}x{outro_height}")

    # 使用 filter_complex 一次性完成：主视频处理 + 统一分辨率 + 拼接
    # concat 前统一视频 fps/像素格式、音频采样率/声道，避免尾部素材参数漂移导致拼接失败。
    # outro 按比例缩放并 pad 到主视频画布，避免 16:9 片尾被拉伸成主视频比例。
    filter_complex = (
        # 主视频处理：加速 + 等比缩放（快速模式）
        f"[0:v]setpts=0.909*PTS,scale={target_width}:{target_height},setsar=1:1,"
        f"fps={FPS},format=yuv420p,"
        "drawtext=text='结尾有惊喜':fontsize=40:fontcolor=white:x=30:y=h-100:"
        "enable='between(t,20,23)'[vout_main];"
        # 主音频处理：加速并规整格式
        "[0:a]atempo=1.1,aformat=sample_rates=44100:channel_layouts=stereo,asetpts=PTS-STARTPTS[aout_main];"
        # 结尾视频等比缩放并补边到与主视频相同分辨率
        f"[1:v]scale={target_width}:{target_height}:force_original_aspect_ratio=decrease,"
        f"pad={target_width}:{target_height}:(ow-iw)/2:(oh-ih)/2,setsar=1:1,"
        f"fps={FPS},format=yuv420p,setpts=PTS-STARTPTS[vout_outro];"
        # 结尾音频规整格式
        "[1:a]aformat=sample_rates=44100:channel_layouts=stereo,asetpts=PTS-STARTPTS[aout_outro];"
        # 拼接视频和音频
        "[vout_main][vout_outro]concat=n=2:v=1:a=0[vout];"
        "[aout_main][aout_outro]concat=n=2:v=0:a=1[aout]"
    )

    cmd = [
        FFMPEG_PATH,
        "-i", input_video,
        "-i", OUTRO_VIDEO,
        "-filter_complex", filter_complex,
        "-map", "[vout]",
        "-map", "[aout]",

        # 视频编码参数
        *build_video_encoding_args(),
        "-r", FPS,

        # 音频编码参数
        "-c:a", AUDIO_CODEC,
        "-b:a", AUDIO_BITRATE,

        # 优化选项
        "-movflags", "+faststart",

        # 覆盖已存在的文件
        "-y",
        temp_output_path
    ]

    print(f"🎬 正在处理视频...")
    print(f"📝 输入: {input_video}")
    print(f"📝 结尾: {OUTRO_VIDEO}")
    print(f"📝 输出: {output_path}")
    print(f"⚙️  参数: 1.1倍速 | {target_width}x{target_height} | {FPS} fps | {VIDEO_CODEC} | 快速模式 | 结尾拼接")

    result = subprocess.run(cmd, capture_output=True)

    if result.returncode != 0:
        if os.path.exists(temp_output_path):
            os.remove(temp_output_path)
        error_msg = result.stderr.decode('utf-8') if result.stderr else result.stdout.decode('utf-8')
        raise RuntimeError(f"❌ ffmpeg 执行失败:\n{error_msg}")

    expected_duration = get_media_duration(input_video) / 1.1 + get_media_duration(OUTRO_VIDEO)
    actual_duration = finalize_output(temp_output_path, output_path, expected_duration)
    print(f"✅ 输出时长校验通过: {actual_duration:.2f}s")

    return output_path


def main():
    lock_file = None
    try:
        lock_file = acquire_process_lock()

        # 1. 校验 ffmpeg
        validate_ffmpeg()

        # 2. 自动获取最新视频（如果命令行指定了路径则使用指定路径）
        if len(sys.argv) >= 2:
            video_path = sys.argv[1]
            print(f"📂 使用指定视频: {video_path}")
        else:
            video_path = get_latest_video()

        # 3. 校验视频
        validate_video(video_path)

        # 4. 处理视频
        output_path = process_video(video_path)

        # 5. 返回结果
        print(f"\n✅ Video2 生成成功: {output_path}")
        print(f"📁 文件位置: {OUTPUT_DIR}")

    except RuntimeError as e:
        print(str(e), file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"❌ 未知错误: {e}", file=sys.stderr)
        sys.exit(1)
    finally:
        if lock_file:
            lock_file.close()


if __name__ == "__main__":
    main()
